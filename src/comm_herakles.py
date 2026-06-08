import os
import re
import time
import signal
import shutil
from pathlib import Path
from datetime import datetime, timedelta
from dotenv import load_dotenv
from pywinauto import Application
from openpyxl import load_workbook
from mailbox_pdf_fetcher import write_log 
from pywinauto.timings import Timings

# -----------------------------------------
# Importation des variables d'environnement
# -----------------------------------------
load_dotenv()

ID_HERAKLES = os.getenv("ID_HERAKLES")
if not ID_HERAKLES:
    raise ValueError("ID_HERAKLES n'est pas définie dans les variables d'environnement.")

MDP_HERAKLES = os.getenv("MDP_HERAKLES")
if not MDP_HERAKLES:
    raise ValueError("MDP_HERAKLES n'est pas définie dans les variables d'environnement.")

FACT_SUPPLIER = os.getenv("FACT_SUPPLIER")
if not FACT_SUPPLIER:
    raise ValueError("FACT_SUPPLIER n'est pas définie dans les variables d'environnement.")

EXE_PATH = os.getenv("EXE_PATH")
if not EXE_PATH:
    raise ValueError("EXE_PATH n'est pas définie dans les variables d'environnement.")

HERAKLES_TRACKER = os.getenv("HERAKLES_TRACKER")
if not HERAKLES_TRACKER:
    raise ValueError("HERAKLES_TRACKER n'est pas définie dans les variables d'environnement.")

Timings.window_find_timeout = 5
Timings.window_find_retry = 0.2   # moins d'appels CPU
# ----------------------------------------
# Processus de communication avec Herakles
# ----------------------------------------

interrupted = False

def handle_sigint(signum, frame):
    global interrupted
    interrupted = True
    write_log(HERAKLES_TRACKER, "Ctrl+C détecté, arrêt après la ligne en cours...")

signal.signal(signal.SIGINT, handle_sigint)

# Connection à l'application
def connection():
    try:
        # On essaie de se connecter si c'est déjà ouvert
        app = Application(backend="uia").connect(path=EXE_PATH, timeout=3)
        write_log(HERAKLES_TRACKER, "Application déjà ouverte, connexion établie.")
    except Exception:
        # Sinon on lance l'appli
        write_log(HERAKLES_TRACKER, "Lancement de l'application...")
        app = Application(backend="uia").start(EXE_PATH)
        time.sleep(1)
    
    window = app.window(title="HERAKLES : Identification de l'utilisateur")
    if window.exists(timeout=20):
        # Permet que les touches du clavier soient au bon endroit
        window.set_focus()

        # Login
        write_log(HERAKLES_TRACKER, "Tentative de saisie par IDs techniques...")
        # Saisie Utilisateur
        combo = window.child_window(auto_id="1", control_type="ComboBox")
        combo.click_input()
        combo.set_focus()
        combo.type_keys("^a{BACKSPACE}" + ID_HERAKLES, pause=0.1)
        
        # Saisie MDP
        pwd = window.child_window(auto_id="3", control_type="Edit")
        pwd.click_input()
        pwd.type_keys(MDP_HERAKLES, pause=0.1)
        
        # Valider
        window.child_window(title="Valider", control_type="Button").click_input()

        # Vérification finale
        time.sleep(1)
        if not window.exists():
            write_log(HERAKLES_TRACKER, "Connexion réussie !")
        else:
            write_log(HERAKLES_TRACKER, "La fenêtre est toujours là. Vérifie manuellement si les champs sont remplis.")

        try:
            main_window = app.window(title_re=".*HERAKLES.*")
            main_window.wait('ready', timeout=10) 
            write_log(HERAKLES_TRACKER, "On est sur la page principale !")
            return main_window
        except:
            write_log(HERAKLES_TRACKER, "Impossible de trouver la fenêtre principale.")
        return None
    else :
        return None

# Placement sur la barre de recherche des commandes fournisseurs
def research_bar(main_window):
    try: 
        # Entre dans la liste des commandes fournisseurs
        write_log(HERAKLES_TRACKER, "Tentative d'entrer dans la liste des commandes fournisseurs")
        achats = main_window.child_window(title="Commandes fournisseurs : toutes", control_type="TreeItem")
        achats.click_input()
        write_log(HERAKLES_TRACKER, "Nous sommes dans la liste des commandes fournisseurs")

        # Clique sur le bouton annuler 
        limit = main_window.child_window(title="Limiter l'affichage ?", control_type="Window")
        if limit.exists(timeout=10):
            annule = limit.child_window(title="Annuler", control_type="Button")
            annule.wait('ready', timeout=5)
        write_log(HERAKLES_TRACKER, "Clic sur annuler")
        annule.click_input()

    except Exception as e:
        write_log(HERAKLES_TRACKER, f"Erreur Herakles à l'étape : {e}")
        raise e


# Sélection de la commande en tapant le numéro dans la barre de recherche
def select_order2(main_window, n_order): 
    try:
        # Entrée du numéro de commande dans la barre de recherche
        recherche = main_window.child_window(auto_id="1001", control_type="Edit")
        recherche.set_focus()
        recherche.click_input()
        recherche.set_text("")
        recherche.type_keys(f"^a{{BACKSPACE}}{n_order}{{ENTER}}", pause=0.1)
        time.sleep(2)
        select = main_window.child_window(title=str(n_order), control_type="DataItem")
        select.wait('ready', timeout=10)
        select.set_focus() 
        select.double_click_input()
        write_log(HERAKLES_TRACKER, f"Nous sommes sur la commande {n_order}")

    except Exception as e:
        write_log(HERAKLES_TRACKER, f"Erreur Herakles à l'étape : {e}")
        raise e
    
def select_order3(main_window, n_order):
    try:
        n_order_str = str(n_order)
     # Caractères de type "tiret" à normaliser
        unicode_dashes = {
            '\u2212': '-',  # MINUS SIGN
            '\u2013': '-',  # EN DASH –
            '\u2014': '-',  # EM DASH —
            '\u2010': '-',  # HYPHEN ‐
            '\u2011': '-',  # NON-BREAKING HYPHEN
            '\u00AD': '-',  # SOFT HYPHEN
        }
        for bad, good in unicode_dashes.items():
            n_order_str = n_order_str.replace(bad, good)
        
        write_log(HERAKLES_TRACKER, f"Recherche de la commande normalisée : {n_order_str}")      
        recherche = main_window.child_window(auto_id="1001", control_type="Edit")
        
        # S'assurer que le champ existe et est prêt
        recherche.wait('ready enabled visible', timeout=10)
        
        # Vider le champ de façon fiable (plus robuste que Ctrl+A+Backspace)
        recherche.set_focus()
        time.sleep(0.3)
        recherche.click_input()
        time.sleep(0.3)
        # set_edit_text est plus fiable que type_keys pour vider
        try:
            recherche.set_edit_text("")
        except Exception:
            recherche.type_keys("^a{BACKSPACE}", pause=0.1)
        time.sleep(0.3)
        
        # Saisir le numéro avec des pauses
        recherche.type_keys(n_order_str, pause=0.08, with_spaces=True)
        time.sleep(0.5)
        recherche.type_keys("{ENTER}")
        
        # Attente active du résultat avec retry (au lieu d'un sleep fixe)
        select = None
        max_attempts = 15  # ~15 secondes max
        for attempt in range(max_attempts):
            time.sleep(1)
            try:
                # Matching plus strict pour éviter les faux positifs
                candidate = main_window.child_window(
                    title=n_order_str,
                    control_type="DataItem"
                )
                if candidate.exists(timeout=0.5):
                    select = candidate
                    break
            except Exception:
                pass
            write_log(HERAKLES_TRACKER, f"Tentative {attempt+1}/{max_attempts} : commande {n_order_str} pas encore visible")
        
        if select is None:
            raise RuntimeError(f"Commande {n_order_str} introuvable après {max_attempts}s d'attente")
        
        # Double-clic avec tentatives
        select.wait('ready visible', timeout=5)
        select.set_focus()
        time.sleep(0.3)
        select.double_click_input()
        write_log(HERAKLES_TRACKER, f"Nous sommes sur la commande {n_order}")

    except Exception as e:
        write_log(HERAKLES_TRACKER, f"Erreur Herakles à l'étape select_order : {e}")
        raise e
    
def select_order(main_window, n_order):
    try:
        n_order_str = str(n_order)
        # Normalisation des tirets Unicode
        unicode_dashes = {
            '\u2212': '-', '\u2013': '-', '\u2014': '-',
            '\u2010': '-', '\u2011': '-', '\u00AD': '-',
        }
        for bad, good in unicode_dashes.items():
            n_order_str = n_order_str.replace(bad, good)

        write_log(HERAKLES_TRACKER, f"Recherche de la commande normalisée : {n_order_str}")

        # On s'assure que la fenêtre principale est bien au premier plan
        # (sinon UIA rame ou échoue si le module .NET est resté devant)
        main_window.wait('exists ready', timeout=15)
        main_window.set_focus()

        recherche = main_window.child_window(auto_id="1001", control_type="Edit")
        recherche.wait('ready enabled visible', timeout=10)

        # Saisie via set_edit_text (instantané, pas de frappe touche par touche)
        recherche.set_focus()
        try:
            recherche.set_edit_text("")
            recherche.set_edit_text(n_order_str)
        except Exception:
            recherche.click_input()
            recherche.type_keys("^a{BACKSPACE}", pause=0.1)
            recherche.type_keys(n_order_str, pause=0.08, with_spaces=True)
        recherche.type_keys("{ENTER}")

        # Attente active du résultat (la grille est peuplée par le serveur, ~lent)
        t_search = time.time()
        select = None
        candidate = main_window.child_window(title=n_order_str, control_type="DataItem")
        max_attempts = 30
        for attempt in range(max_attempts):
            if candidate.exists(timeout=1):
                select = candidate
                write_log(HERAKLES_TRACKER,
                          f"Commande visible après {time.time()-t_search:.1f}s "
                          f"(tentative {attempt+1})")
                break
            # pas de sleep ici : exists(timeout=1) fait déjà l'attente d'1s

        if select is None:
            raise RuntimeError(
                f"Commande {n_order_str} introuvable après {time.time()-t_search:.1f}s")

        # Double-clic, instrumenté
        t_click = time.time()
        select.wait('ready visible', timeout=10)
        select.set_focus()
        select.double_click_input()
        write_log(HERAKLES_TRACKER,
                  f"Double-clic effectué en {time.time()-t_click:.1f}s")
        write_log(HERAKLES_TRACKER, f"Nous sommes sur la commande {n_order}")

    except Exception as e:
        write_log(HERAKLES_TRACKER, f"Erreur Herakles à l'étape select_order : {e}")
        raise e

# Connexion à la nouvelle fenêtre des Achats .NET via son titre 
def get_win_achats():
    try:
        write_log(HERAKLES_TRACKER, "Tentative de liaison directe via le titre...")
        app_achats = Application(backend="uia").connect(title_re=".*Achats - serveur-2016.*", timeout=5)
        win_achats = app_achats.window(title_re=".*Achats - serveur-2016.*")
        write_log(HERAKLES_TRACKER, "Succès ! On est connecté au module .NET.")
        return win_achats
    
    except Exception as e:
        write_log(HERAKLES_TRACKER, f"Erreur Herakles à l'étape : {e}")
        raise e

# Clic sur le bouton pour accéder aux liens
def clic_links(win_achats):
    try:
        write_log(HERAKLES_TRACKER, "Tentative de clic sur le bouton des liens")
        links = win_achats.child_window(title="Liens", control_type="Button")
        links.click_input()
        write_log(HERAKLES_TRACKER, "Succès ! On a accès aux liens.")

    except Exception as e:
        write_log(HERAKLES_TRACKER, f"Erreur Herakles à l'étape : {e}")
        raise e

# Recherche si la facture FF a déjà été créée
def try_ff(win_achats):
    fact = win_achats.child_window(title_re=r"FF2[56]-.*", control_type="DataItem")
    if not fact.exists(timeout=3):
        return None
    num_ff = fact.window_text()
    write_log(HERAKLES_TRACKER, "Le numéro de FF existe déjà")
    return num_ff

def get_brf2(win_achats):
    regex = re.compile(r"BRF2[56]-")
    for _ in range(10):
        all_items = win_achats.descendants(control_type="DataItem")
        liens = [it for it in all_items if regex.search(it.window_text() or "")]
        if liens:
            write_log(HERAKLES_TRACKER, f"{len(liens)} BRF trouvés. Début du traitement...")
            return liens
        time.sleep(0.4)
    write_log(HERAKLES_TRACKER, "Aucun BRF trouvé")
    return None

def get_brf(win_achats):
    regex = re.compile(r"BRF2[56]-")
    t0 = time.time()
    for attempt in range(10):
        all_items = win_achats.descendants(control_type="DataItem")
        liens = [it for it in all_items if regex.search(it.window_text() or "")]
        if liens:
            write_log(HERAKLES_TRACKER,
                      f"{len(liens)} BRF trouvés en {time.time()-t0:.1f}s "
                      f"(tentative {attempt+1}, {len(all_items)} DataItems scannés)")
            return liens
        if attempt == 0:
            # au 1er tour, on logue combien d'items existent pour diagnostic
            write_log(HERAKLES_TRACKER,
                      f"Tour 1 : {len(all_items)} DataItems présents, aucun BRF parmi eux")
        time.sleep(0.4)
    write_log(HERAKLES_TRACKER, f"Aucun BRF trouvé après {time.time()-t0:.1f}s")
    return None

# Fonction qui vérifie si le BRF est là et clique dessus si oui
def try_brf(lien) : 
    lien.click_input()
    write_log(HERAKLES_TRACKER, "Succès ! On a cliqué sur le bon de récéption")
    return True

# Fonction qui entre le numéro de facture dans la ref fact fournisseur 
def try_n_fact(win_achats, n_fact):
    try:
        ref_fact = win_achats.child_window(title="Réf. fact. fourn.", control_type="Edit")
        ref_fact.wait('ready enabled visible', timeout=10)
        try:
            ref_fact.set_edit_text(str(n_fact))   # instantané, pas de frappe
        except Exception:
            # Repli si le champ refuse set_edit_text (contrôle custom)
            ref_fact.set_focus()
            time.sleep(0.3)
            ref_fact.type_keys("^a{BACKSPACE}", pause=0.1)
            ref_fact.type_keys(str(n_fact), with_spaces=True, pause=0.1)
        ref_fact.type_keys("{ENTER}")
       # ref_fact.type_keys(f"{str(n_fact)}", with_spaces=True, pause=0.2)
        #time.sleep(0.5)
        #ref_fact.type_keys("{ENTER}")
    except Exception as e:
        write_log(HERAKLES_TRACKER, f"Erreur Herakles à l'étape : {e}")
        raise e


# Récupération du montant HT
def get_ht_field(win_achats):
    try:
        # Vérification du montant total
        write_log(HERAKLES_TRACKER, "On tente de vérifier le montant HT")
        ht_field = win_achats.child_window(title="Montant HT", control_type="Edit", found_index=0)
        write_log(HERAKLES_TRACKER, "On récupère le montant HT.")                                                     
        # Vérification du prix
        valeur_affichee = ht_field.get_value() 
        write_log(HERAKLES_TRACKER, f"Montant récupéré : {valeur_affichee}")

        clean_string = re.sub(r'[^\d,.]', '', valeur_affichee).replace(',', '.')
        price_float = float(clean_string)
        write_log(HERAKLES_TRACKER, f"Montant nettoyé : {price_float}")
        return price_float
    except Exception as e:
        write_log(HERAKLES_TRACKER, f"Erreur Herakles à l'étape : {e}")
        raise e
    
# Sauvegarde de la facture
def save_fact(win_achats, ht_field):
    try:
        save = win_achats.child_window(title="Enregistrer", control_type="Button")
        save.click_input()
        same_n_fact = win_achats.child_window(title_re=".*Achat-serveur.*", control_type="Window")
        if not same_n_fact.exists(timeout=2):
            same_n_fact = win_achats.child_window(control_type="Window", top_level_only=False, found_index=0)
        if same_n_fact.exists():
            write_log(HERAKLES_TRACKER, "Fenêtre d'alerte détectée. Tentative de clic sur 'Oui'.")
            yes = same_n_fact.child_window(title="Oui", control_type="Button")
            if yes.exists(timeout=3) :
                yes.set_focus()
                yes.click_input()
        created_ff_text = None
        win_save = win_achats.child_window(title="Enregistrement facture fournisseur", control_type="Window")

        if win_save.exists(timeout=10):
            field = win_save.child_window(title_re="FF2[56]-.*")
            if field.exists(timeout=2):
                created_ff_text = field.window_text()
                write_log(HERAKLES_TRACKER, f"Facture {created_ff_text} créé !")
            write_log(HERAKLES_TRACKER, "Clic sur le premier valider pour enregistrer la fact")
            first_valid = win_save.child_window(title="Valider", control_type="Button")
            first_valid.wait('ready', timeout=5)
            first_valid.click_input()
    
        # Si le montant HT est < à 2500 alors on valide la facture 
        if ht_field < 2500:
            write_log(HERAKLES_TRACKER, "Le montant est < à 2500 :")
            write_log(HERAKLES_TRACKER, "Clic sur valider de fact final")
            valid = win_achats.child_window(title="Valider", control_type="Button")
            valid.wait('ready', timeout=5)
            valid.click_input()

        return created_ff_text
    except Exception as e:
        write_log(HERAKLES_TRACKER, f"Erreur Herakles à l'étape : {e}")
        raise e

# Mécanisme de création de facture 
def create_fact(win_achats, n_fact):
    try:
        # Clic sur le bouton pour créer la facture
        write_log(HERAKLES_TRACKER, "Tentative de clic sur le bouton de création de facture")
        crea_fact = win_achats.child_window(title="Créer facture", control_type="Button")
        crea_fact.click_input()
        write_log(HERAKLES_TRACKER, "Succès ! On peut commencer à créer la facture")

        try_n_fact(win_achats, n_fact)
        ht_field = get_ht_field(win_achats)
        created_ff = save_fact(win_achats, ht_field)
        return created_ff
    except Exception as e:
        write_log(HERAKLES_TRACKER, f"Erreur Herakles à l'étape : {e}")
        return None
    
# Gestion des jours de la semaine
def get_lookback_date():
    today = datetime.now().date()
    weekday = today.weekday()  # Lundi=0, Mardi=1, Mercredi=2, Jeudi=3, Vendredi=4

    # Si on est Mercredi (2) ou Vendredi (4), on remonte de 2 jours
    if weekday == 2:
        days_to_subtract = 4
    elif weekday == 4:
        days_to_subtract = 0
    # Par défaut, on remonte de 1 jour
    else:
        days_to_subtract = 1
    return today - timedelta(days=days_to_subtract)

def safe_close(window):
    """Ferme une fenêtre pywinauto en ignorant les erreurs si elle n'existe plus."""
    if window is None:
        return
    try:
        if window.exists():
            window.close()
    except Exception:
        # La fenêtre a déjà disparu ou est inaccessible, on ignore
        pass

def safe_save(wb, path):
    """Sauvegarde atomique avec backup."""
    path = Path(path)
    backup = path.with_suffix('.xlsx.bak')
    tmp = path.with_suffix('.xlsx.tmp')
    
    # 1. Sauvegarde dans un fichier temporaire
    wb.save(tmp)
    # 2. Backup de l'ancien fichier s'il existe
    if path.exists():
        shutil.copy2(path, backup)
    # 3. Remplacement atomique
    os.replace(tmp, path)

# Mise à jour du tableau excel en fonction du résultat d'herakles et exécution des actions sur herakles
def update_xlsx(): 

    tmp_file = Path(FACT_SUPPLIER).with_suffix('.xlsx.tmp')
    if tmp_file.exists():
        write_log(HERAKLES_TRACKER, f"Fichier temporaire résiduel détecté, suppression : {tmp_file}")
        try:
            tmp_file.unlink()
        except Exception as e:
            write_log(HERAKLES_TRACKER, f"Impossible de supprimer le .tmp résiduel : {e}")

    wb = load_workbook(FACT_SUPPLIER)
    ws = wb.active
    col_map = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        col_map[cell.value] = col_idx

    #start_date = get_lookback_date()
    #today = datetime.now().date()
    precise_date = datetime(2026, 6, 4, 0, 0, 0).date()
    write_log(HERAKLES_TRACKER, "========================================")
    #write_log(HERAKLES_TRACKER, f"Analyse des dates du {precise_date} au {today}")
    write_log(HERAKLES_TRACKER, f"Analyse des dates au {precise_date}")
    main_window = connection()
    if main_window is None:
        write_log(HERAKLES_TRACKER, "ARRÊT CRITIQUE : Impossible d'ouvrir Herakles. Le script s'arrête ici.")
        exit()
    research_bar(main_window)

    try :
    # On boucle sur la date du jour
        for row_idx in range(2, ws.max_row + 1):
            if interrupted :
                write_log(HERAKLES_TRACKER, f"Interruption demandée, arrêt à la ligne {row_idx}.")
                break
            date_val = ws.cell(row=row_idx, column=col_map["DATE TRAITEMENT"]).value
            # On vérifie que date_val est non vide et que la valeur est bien sous format de date et on filtre sur JJ/MM/AAAA

            # Pour filtrer d'une date à une autre
            # if not (date_val and hasattr(date_val, 'date') and precise_date <= date_val.date() <= today):

            # Pour filtrer sur une date précise 
            if not (date_val and hasattr(date_val, 'date') and date_val.date() == precise_date):
                continue
            write_log(HERAKLES_TRACKER, f"Lancement sur la date du {date_val}")
            n_order = ws.cell(row=row_idx, column=col_map["N COMMANDE"]).value
            write_log(HERAKLES_TRACKER, f"Numéro de commande CF : {n_order}")
            n_fact = ws.cell(row=row_idx, column=col_map["CODE FACTURE"]).value
            write_log(HERAKLES_TRACKER, f"Numéro de facture : {n_fact}")
            win_achats = None

            try:
                # Si on a pas de CF -> message dans excel et on passe à la ligne suivante
                if not n_order:
                    write_log(HERAKLES_TRACKER, "n_order est non trouvable -> on passe à la facture suivante")
                    ws.cell(row=row_idx, column=col_map["MESSAGE HERAKLES"], value="Absence : n° de commande CF.")
                    continue
                
                # Si on n'a pas de n° de facture -> message dans excel et on passe à la ligne suivante
                if not n_fact:
                    write_log(HERAKLES_TRACKER, "n_fact est non trouvable -> on passe à la facture suivante")
                    ws.cell(row=row_idx, column=col_map["MESSAGE HERAKLES"], value="Absence : n° de facture.")
                    continue
                
                # Si la facture a déjà été créée
                if ws.cell(row=row_idx, column=col_map["HERAKLES"]).value == "oui" :
                    write_log(HERAKLES_TRACKER, f"La facture associée à la commande {n_order} existe déjà.")
                    continue
            
                # Si on a tout -> lancement du script en entier
                write_log(HERAKLES_TRACKER, f"Lancement du traitement de la commande: {n_order}")
                select_order(main_window, n_order)
                win_achats = get_win_achats()
                clic_links(win_achats)
                num_ff = try_ff(win_achats)
                if num_ff is None:
                    # Si le brf est absent, on met le message et on passe au prochain numéro dans la colonne
                    write_log(HERAKLES_TRACKER, "Recherche du ou des BRF")
                    liens = get_brf(win_achats)
                    if liens is None :
                        ws.cell(row=row_idx, column=col_map["MESSAGE HERAKLES"], value="Le BRF est manquant.")
                        safe_close(win_achats)
                        continue
                    for lien in liens :
                        try_brf(lien)
                        write_log(HERAKLES_TRACKER, "On tente de lancer la création de facture")
                        created_ff = create_fact(win_achats, n_fact)
                        if created_ff :
                            ws.cell(row=row_idx, column=col_map["HERAKLES"], value="oui")

                            # Si on a déjà un message, on ajoute le nouveau à la suite (cas plusieurs BRF)
                            current_value = ws.cell(row=row_idx, column=col_map["MESSAGE HERAKLES"]).value or ""
                            new_message = f"Facture {created_ff} créée !"
                            if current_value:
                                updated_value = f"{current_value} | {new_message}"
                            else:
                                updated_value = new_message
                            ws.cell(row=row_idx, column=col_map["MESSAGE HERAKLES"], value=updated_value)
                            #ws.cell(row=row_idx, column=col_map["MESSAGE HERAKLES"], value=f"Facture {created_ff} créée !")

                            write_log(HERAKLES_TRACKER, f"Facture {created_ff} pour la commande {n_order} créé avec succès !")

                            # Si on a déjà un montant HT, on ajoute le nouveau à la suite (cas plusieurs BRF)
                            ht_field = get_ht_field(win_achats)
                            current_ht = ws.cell(row=row_idx, column=col_map["MONTANT HT"]).value or ""
                            new_ht = f"{ht_field}"
                            if current_ht:
                                updated_ht = f"{current_ht} | {new_ht}"
                            else:
                                updated_ht = new_ht
                            ws.cell(row=row_idx, column=col_map["MONTANT HT"], value=updated_ht)
                            
                            if ht_field >= 2500 :
                                ws.cell(row=row_idx, column=col_map["A VALIDER"], value="en attente")
                                write_log(HERAKLES_TRACKER, f"Facture de {n_order} à valider car le montant HT est >= 2500.")
                            else :
                                write_log(HERAKLES_TRACKER, f"La facture de {n_order} n'a pas besoin d'être validée car le montant HT est < 2500.")
                else :
                    safe_close(win_achats)
                    write_log(HERAKLES_TRACKER, f"La facture {num_ff} existe déjà")
                    ws.cell(row=row_idx, column=col_map["HERAKLES"], value="oui")
                    ws.cell(row=row_idx, column=col_map["MESSAGE HERAKLES"], value=f"La facture {num_ff} existe déjà.")
                    continue
            except Exception as error:
                write_log(HERAKLES_TRACKER, f"Erreur ligne {row_idx} ({n_order}): {error}")
                if ws.cell(row=row_idx, column=col_map["MESSAGE HERAKLES"]).value is None:
                    ws.cell(row=row_idx, column=col_map["MESSAGE HERAKLES"], value=str(error))
            finally:
                safe_close(win_achats)

                # Sauvegarde incrémentale avec retry en cas de verrou Excel
                saved = False
                for attempt in range(5):
                    try:
                        safe_save(wb, FACT_SUPPLIER)
                        saved = True
                        break
                    except PermissionError:
                        if attempt < 4:
                            write_log(HERAKLES_TRACKER, f"Fichier Excel verrouillé, réessai {attempt+1}/5 dans 3s...")
                            time.sleep(3)
                    except Exception as e:
                        write_log(HERAKLES_TRACKER, f"Erreur sauvegarde ligne {row_idx} : {e}")
                        break
                if not saved:
                    write_log(HERAKLES_TRACKER, f"ATTENTION : ligne {row_idx} non sauvegardée sur disque (en mémoire uniquement)")
    finally:
        # Sauvegarde finale sécurisée
        try:
            safe_save(wb, FACT_SUPPLIER)
            write_log(HERAKLES_TRACKER, "Fichier sauvegardé et fermé proprement.")
        except Exception as e:
            write_log(HERAKLES_TRACKER, f"ERREUR FATALE SAUVEGARDE FINALE : {e}")
        
        safe_close(main_window)

    write_log(HERAKLES_TRACKER, "Processus terminé.")
    write_log(HERAKLES_TRACKER, "========================================\n")

if __name__ == "__main__" :
    update_xlsx()
