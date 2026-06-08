import io
import re
import os
import time
import fitz
import json
import shutil
import pdfplumber
import pytesseract
from PIL import Image
from mistralai import Mistral
from datetime import datetime
from dotenv import load_dotenv
from openpyxl import load_workbook
from rapidfuzz import process, fuzz
from openpyxl.styles import PatternFill
from mailbox_pdf_fetcher import write_log, download_attachments


# --------------------------------------------------------------------
# Importation des variables d'environnement et définition de variables
# --------------------------------------------------------------------

load_dotenv()

DOWNLOAD_FOLDER = os.getenv("DOWNLOAD_FOLDER")
if not DOWNLOAD_FOLDER:
    raise ValueError("DOWNLOAD_FOLDER n'est pas définie dans les variables d'environnement.")

PROCESSED_FOLDER = os.getenv("PROCESSED_FOLDER")
if not PROCESSED_FOLDER:
    raise ValueError("PROCESSED_FOLDER n'est pas définie dans les variables d'environnement.")

TO_CHECK_FOLDER = os.getenv("TO_CHECK_FOLDER")
if not TO_CHECK_FOLDER:
    raise ValueError("TO_CHECK_FOLDER n'est pas définie dans les variables d'environnement.")

TO_FOLLOW_UP = os.getenv("TO_FOLLOW_UP")
if not TO_FOLLOW_UP:
    raise ValueError("TO_FOLLOW_UP n'est pas définie dans les variables d'environnement.")

FILE_LOCATION_TRACKER = os.getenv("FILE_LOCATION_TRACKER")
if not FILE_LOCATION_TRACKER:
    raise ValueError("FILE_LOCATION_TRACKER n'est pas définie dans les variables d'environnement.")

FACT_SUPPLIER = os.getenv("FACT_SUPPLIER")
if not FACT_SUPPLIER:
    raise ValueError("FACT_SUPPLIER n'est pas définie dans les variables d'environnement.")

SUPPLIER_CODES = os.getenv("SUPPLIER_CODES")
if not SUPPLIER_CODES:
    raise ValueError("SUPPLIER_CODES n'est pas définie dans les variables d'environnement.")

# Pour la connexion à Mistral
API_KEY = os.getenv("MISTRAL_API_KEY")
MODEL = "mistral-small-latest"
if not API_KEY:
    raise ValueError("La clé API MISTRAL_API_KEY n'est pas définie dans les variables d'environnement.")

client = Mistral(api_key=API_KEY)


# Défintions des valeurs invalides 
INVALID_VALUES = re.compile(r"\b(TVA|IBAN|RIB|SIRET|CRCA|ÉMETTEUR|CLIENT|ÉMISSION|DATE|PAIEMENT)\b[A-Z0-9\-_\/]*", re.IGNORECASE)
INVALID_SUPPLIER_NAME = re.compile(r"\b(TVA|IBAN|RIB|Relevé d’Identité Bancaire|SIRET|CRCA|FR25|ÉMETTEUR|CLIENT|EBENE ET TRADITION|EBENE|WOOD EN|PAGE|FACTURE|NUMÉRO|DÉBITEUR|COPIEPRIVEE|COPIEFRANCE|CGV|CONDITIONS GÉNÉRALES DE VENTE|COMPTE|DATE|CODE|RÉGLEMENT)\s*([^,]+)*[A-Z0-9\-_\/]*", re.IGNORECASE)


# ------------------------------------
# Lecture des factures pdf ou scannées
# ------------------------------------

# Lecture du pdf et le renvoie sous forme de texte
def read_pdf(pdf_path):
    text = ''
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + '\n'
        return text

# Conversion du pdf en image et le renvoie sous forme de texte
def pdf_to_jpeg(pdf_path):
    try:
        doc = fitz.open(pdf_path)
        full_text = ''
        for page_index in range(len(doc)):
            page = doc[page_index]
            
            # Conversion de la page en image (zoom 2x pour une meilleure précision OCR)
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
            img_data = pix.tobytes("png")
            img = Image.open(io.BytesIO(img_data))
            
            # OCR avec Tesseract sur l'image générée
            page_text = pytesseract.image_to_string(img, lang='fra')
            
            if page_text:
                full_text += page_text + '\n'
        
        doc.close()
        return full_text
        
    except Exception as e:
        print(f"Erreur lors de l'extraction (PyMuPDF + Tesseract) : {e}")
        return ""
    

# ---------------------------------------------
# Fonctions pour vérifier la validité des infos 
# ---------------------------------------------

# Vérifie si le numéro récupéré est une date
def is_date(tok):
    return bool(re.match(r"\d{2}[/-]\d{2}[/-]\d{4}$|\d{4}[/-]\d{2}[/-]\d{2}$|\d{1,2}\s*(?:janvier|février|mars|avril|mai|juin|juillet|août|septembre|octobre|novembre|décembre)\s*\d{4}$", tok))

# Vérifie si le numéro récupéré est un code postal
def is_postal_code(tok):
    return bool(re.match(r"\b(?:0[1-9]|[1-8][0-9]|9[0-5]|2A|2B)\d{3}(?:\s+[A-ZÉÈA-Z\-]+)?\b", tok))

# Vérifie si le numéro récupéré est une adresse
def is_adress(tok):
    #return bool(re.search(r"[0-9]*(?:rue|impasse|avenue)(\w+)*", tok, re.IGNORECASE))
    return bool(re.search(r"[0-9]*(?:(?:[,\s]+(?:bis|ter|quater)?[,\s]+)?(?:rue|avenue|boulevard|chemin|allée|impasse|place|square|quai|route|lotissement|résidence|passage|ville|citée|voie)[,\s]+[A-Za-zÀ-ÿ\s\-']+)\b", tok, re.IGNORECASE))

# Vérifie si le nom récupéré est valide
def is_valid_value(s, text, extracted_num_fact):
    return (not is_date(s) and not is_postal_code(s) and not is_adress(s) and not is_n_fact(s, extracted_num_fact) and not get_num_order(text))

# Vérifie si l'argument est une adresse mail
def is_email_adress(tok) :
    return bool(re.search(r"([^\s:@]+)@([^\s:@.]+)", tok, re.IGNORECASE))

# Vérifie si le numéro récupéré est un numéro de facture
def is_n_fact(tok, extracted_num_fact):
    if extracted_num_fact and extracted_num_fact != "None":
        return tok == extracted_num_fact
    return False


# --------------------
# Extraction des infos
# --------------------

# Communication avec Mistral pour extraire certaines infos
def get_invoice_data_ia(text):
    """Effectue un seul appel IA pour récupérer toutes les infos d'un coup."""
    prompt = f"""
    Tu es un expert en lecture de factures. Analyse le texte suivant et renvoie UNIQUEMENT un objet JSON avec ces clés :
    - "num_fact": Un numéro de facture DOIT être une chaîne alphanumérique de 2 à 20 caractères. un numéro de facture ne peut PAS commencer par CF25- ou CF26-, sinon "None"
    - "dates": une liste de chaînes ["date_facture", "date_echeance"] (ex: ["20/01/2025", "15/02/2025"])
    - "total_ttc": le montant total TTC
    
    Texte de la facture :
    {text}
    """
    try:
        response = client.chat.complete(
            model=MODEL,
            messages=[{"role": "user", "content": prompt}],
            response_format={"type": "json_object"}
        )
        return json.loads(response.choices[0].message.content)
    except Exception as e:
        print(f"Erreur API Mistral : {e}")
        return None
# Extraction du numéro de commande CF25-XXXX
def get_num_order2(text):
    matches = re.findall(r"\bCF[\s-]*2[56]\s*-\s*[0-9]{4}\b", text, re.IGNORECASE)
    cleaned_matches = []
    for m in matches :
        if m:
            cleaned_matches.append(m.replace(" ", "").replace("CF-2", "CF2").replace("CF 026-", "CF26-"))
    return cleaned_matches

def get_num_order3(text):
    matches = re.findall(r"CF[\s-]*(2[56])[\s-]*([0-9]{4})\b", text)
    return [f"CF{m[0]}-{m[1]}" for m in matches]

_SEP = r"[-\u2212\u2013\u2014\u2010\u2011\u00AD\s]*"

def get_num_order(text):
    pattern = rf"CF{_SEP}0?{_SEP}(2[56]){_SEP}([0-9]{{4}})\b"
    matches = re.findall(pattern, text, re.IGNORECASE)
    return [f"CF{m[0]}-{m[1]}" for m in matches]


# Corrige la date si l'année n'est pas notée en entier
def corrected_year(date_str, fmt):

    # Si l'année est placé en premier et que c'est 25 et non 2025 -> on la note en entier
    if fmt.startswith('%Y'):
        match = re.match(r'^(25)[/\-\.\s]', date_str)
        if match:
            year = match.group(1)
            year_complete = "20" + year
            return date_str.replace(year, year_complete)
        
    # Si l'année est placé en dernier et que c'est 25 et non 2025 -> on la note en entier
    if fmt.endswith('%Y'):
        match = re.search(r'[/\-\.\s](25)$', date_str)
        if match:
            year = match.group(1)
            year_complete = "20" + year
            return date_str.replace(year, year_complete)
        
    return date_str

# Parse la date en objet datetime (tous les formats sont acceptés)
def parse_date(date_str):

    date_str = date_str.strip()
    date_str = re.sub(r"[^0-9a-zA-Zéèêàçùûôïü\s/.-]", "", date_str)

    MONTHS_MAP = {
        "janvier": 1, "février": 2, "mars": 3, "avril": 4, "mai": 5, "juin": 6,
        "juillet": 7, "août": 8, "septembre": 9, "octobre": 10, "novembre": 11, "décembre": 12
    }
    
    # Convertit les dates sous format classique en chiffres
    for fmt in ('%Y/%m/%d', '%Y-%m-%d', '%Y.%m.%d', '%Y %m %d', '%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y', '%d %m %Y'):
        corrected_date = corrected_year(date_str, fmt)
        try:
            date = datetime.strptime(corrected_date, fmt)
            if date.year >= 2025:
                return date
            else:
                return None
        except ValueError:
            continue

    # Convertit les dates sous format avec le mois écrit en entier
    month_pattern = r"\d{1,2}\s*(?:janvier|février|mars|avril|mai|juin|juillet|août|septembre|octobre|novembre|décembre)\s*\d{4}"
    if re.fullmatch(month_pattern, date_str, re.IGNORECASE):
        parts = re.split(r'\s+', date_str, maxsplit=2)
        if len(parts) == 3:
            day, month_name, year = parts
            month_name_lower = month_name.lower()
            if month_name_lower in MONTHS_MAP:
                month_num = MONTHS_MAP[month_name_lower]
                try:
                    date = datetime(int(year), month_num, int(day))
                    if date.year >= 2025:
                        return date 
                except ValueError:
                    pass

    return None

# Prend les dates renvoyées par l'IA et les trie dans l'ordre chronologique et supprime les doublons avant de les renvoyer
def get_dates(lst_dates_raw):

    if not lst_dates_raw or not isinstance(lst_dates_raw, list):
        return [None, None]

    parsed = []
    for d_str in lst_dates_raw:
        dt = parse_date(str(d_str))
        if dt: parsed.append((str(d_str), dt))
    
    sorted_res = [d[0] for d in sorted(set(parsed), key=lambda x: x[1])]
    return [sorted_res[0] if len(sorted_res) >= 1 else None, 
            sorted_res[1] if len(sorted_res) >= 2 else None]


# Extraction du code fournisseur dans le fichier excel
def get_code_supplier(supplier):
    
    if supplier is None :
        return None

    print(f"---------{supplier}------------")
    
    supplier = supplier.upper()
    # Ouverture d'un fichier xlsx existant 
    wb = load_workbook(SUPPLIER_CODES)
    ws = wb.active

    # Créer directement un dictionnaire {code: nom}
    fournisseurs = {}
    for row in ws.iter_rows(min_row=7, max_row=583, min_col=1, max_col=2):
        code_cell, nom_cell = row[0].value, row[1].value  # A = code, B = nom
        if code_cell and nom_cell:
            fournisseurs[str(nom_cell).upper()] = str(code_cell)

    if not fournisseurs:
        return None

    # Recherche du nom du fournisseur (>= 55% de match)
    best_res = process.extractOne(supplier, fournisseurs.keys(), scorer=fuzz.ratio)
    if not best_res or (best_res[1] < 55):
        return None
    best_name = best_res[0]

    return fournisseurs[best_name]


# Importation des données dans un fichier excel
def import_into_xlsx(fact_num, date_facture, supplier_code, date_echeance, num_order):
    
    match_supplier = re.fullmatch(r"[A-Z]{3}[0-9]{3}", supplier_code or "")
    is_valid = True
    if fact_num == "None" or fact_num is None or match_supplier is None:
        is_valid = False

    # Si aucune info n'a pu être récupérée, on renvoie false et on écrit l'erreur dans le fichier log 
    if all(v in ["None", "", None] for v in [fact_num, date_facture, supplier_code, date_echeance, num_order]):
        write_log(FILE_LOCATION_TRACKER, "[ERREUR] Aucune info exploitable, import impossible.")
        write_log(FILE_LOCATION_TRACKER, "=========================================================================\n")
        return False
        
    # Ouverture d'un fichier xlsx existant 
    wb = load_workbook(FACT_SUPPLIER)
    ws = wb.active

    col_date_idx = None
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value == "DATE TRAITEMENT":
            col_date_idx = col_idx
            break

    # On part de la ligne 2 (sous l'entête) et on descend jusqu'à trouver une case vide
    last_row = 2
    while ws.cell(row=last_row, column=col_date_idx).value is not None:
        last_row += 1

    date = datetime.now().date()
    # Dictionnaire pour mapper les colonnes par nom
    col_map = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        col_map[cell.value] = col_idx

    # Ecriture des autres données dans la copie 
    ws.cell(row=last_row, column=col_map["DATE"], value=date_facture)
    ws.cell(row=last_row, column=col_map["DATE ECHEANCE"], value=date_echeance)
    ws.cell(row=last_row, column=col_map["DATE TRAITEMENT"], value=date)

    # Si le CF n'est pas présent, on colore la cellule mais on laisse dans traite
    cell = ws.cell(row=last_row, column=col_map["N COMMANDE"], value=num_order)
    if num_order is None :
        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Si le numéro de facture n'a pas pu être lu, on colore la cellule en rouge et on renvoie False 
    cell = ws.cell(row=last_row, column=col_map["CODE FACTURE"], value=fact_num)
    if fact_num == "None":
        cell.value = "N_fact_incorrect"
        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Si le code fournisseur n'a pas pu être lu, on colore la cellule en orange et on renvoie False 
    cell = ws.cell(row=last_row, column=col_map["CODE FOURNISSEUR"], value=supplier_code)
    if match_supplier is None :
        cell.value = "Code_fourni_absent"
        cell.fill = PatternFill(start_color="FFC880", end_color="FFC880", fill_type="solid")

    # Sauvegarde du fichier en réécrivant le fichier avec les données
    wb.save(FACT_SUPPLIER)

    return is_valid

# Traitement de tous les pdf
def process_all_pdfs():
    files_to_process = download_attachments() 
    results = []

    for item in files_to_process:
        file = item["filename"]       
        pdf_path = item["filepath"]   
        supplier = item["sender_name"] 

        # On essaie de lire en pdf en premier, si ça échoue -> lecture sous format image
        text = read_pdf(pdf_path)
        if text == '':
            text = pdf_to_jpeg(pdf_path)
        
        data_ia = get_invoice_data_ia(text)
        time.sleep(3)

        # Extraction des infos
        if data_ia:
            fact_num = str(data_ia.get("num_fact", "None"))
            print(f"NUM DE FACT : {fact_num}")
            date_facture, date_echeance = get_dates(data_ia.get("dates", []))
        else:
            fact_num, date_facture, date_echeance = "None", None, None

        list_orders = get_num_order(text)
        supplier_code = get_code_supplier(supplier)

        # La source est déjà connue (pdf_path)
        source = pdf_path 
        success = True

        if not list_orders :
            import_into_xlsx(fact_num, date_facture, supplier_code, date_echeance, None)
            destination = os.path.join(TO_FOLLOW_UP, file)
        # On boucle sur nos tâches (chaque tâche = une ligne dans Excel)
        # Si l'import a échoué (False), le pdf est placé dans le dossier à vérifier 
        else :
            for cf_value in list_orders:
                if not import_into_xlsx(fact_num, date_facture, supplier_code, date_echeance, cf_value):
                    success = False
            if success:
                destination = os.path.join(PROCESSED_FOLDER, file)
            else:
                destination = os.path.join(TO_CHECK_FOLDER, file)

        # Déplacement physique du fichier
        if os.path.exists(source):
            if os.path.exists(destination):
                os.remove(destination)
            shutil.move(source, destination)

        # Stockage des résultats
        results.append({
            "file": file,
            "num_fact": fact_num,
            "supplier": supplier,
            "supplier_code": supplier_code
        })

        folder = os.path.basename(os.path.dirname(destination))
        write_log(FILE_LOCATION_TRACKER, f"[INFO] {file} -> Supplier: {supplier}, Folder: {folder}")
        write_log(FILE_LOCATION_TRACKER, "=========================================================================\n")

    return results

if __name__ == "__main__":
    process_all_pdfs()